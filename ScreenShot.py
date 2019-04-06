#coding:utf-8

from selenium import webdriver
from PIL import ImageGrab
from pymouse import PyMouse
import openpyxl
import win32api
import win32gui
import win32print
import win32con
import win32ui
import os
import time

def Time(driver,first_column_content):
	fetchStart = driver.execute_script("return window.performance.timing.fetchStart")
	loadEventEnd = driver.execute_script("return window.performance.timing.loadEventEnd")
	finishTime = (loadEventEnd - fetchStart) / 1000
	print("{}:{}".format(first_column_content,finishTime))
	return finishTime

if __name__ == "__main__":
	chrome_options = webdriver.ChromeOptions()
	m = PyMouse()
	file_name = r'test.xlsx'
	inwb = openpyxl.load_workbook(file_name)
	sheetnames = inwb.get_sheet_names()
	ws = inwb.get_sheet_by_name(sheetnames[0])
	max_rows = ws.max_row
	max_cols = ws.max_column
	for row_number in range(1,(max_rows+1)):
		driver = webdriver.Chrome(executable_path=r'C:\Program Files (x86)\Google\Chrome\Application\chromedriver.exe',chrome_options=chrome_options)
		driver.maximize_window()
		win32api.keybd_event(123, 0, 0, 0)
		time.sleep(1)
		m.click(1668, 124, 1)
		time.sleep(1)
		x1, y1 = 1876, 127
		m.move(x1, y1)
		win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, x1, y1, 0, 0)
		time.sleep(1)
		x2, y2 = 1865, 161
		m.click(x2, y2, 1)
		time.sleep(1)
		first_column_content = ws.cell(row_number, 1).value
		second_column_content = ws.cell(row_number, 2).value
		driver.get(second_column_content)
		fun = Time(driver,first_column_content)
		if fun <= 5:
			time.sleep(fun)
		elif 5<fun<=10:
			time.sleep(fun - 5)
		elif 10<fun<=20:
			time.sleep(fun-10)
		elif 20 < fun <= 30:
			time.sleep(fun - 20)
		elif 30 < fun <= 40:
			time.sleep(fun - 30)
		elif 40 < fun <= 50:
			time.sleep(fun - 40)
		elif 50< fun <= 60:
			time.sleep(fun-50)
		else:
			time.sleep(fun-60)
		im = ImageGrab.grab()
		im.save("C:/Users/GodOfWar/Desktop/python3/网页load时间加载/{}.png".format(first_column_content))
		driver.close()


